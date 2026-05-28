"""
tests/test_excel_export.py
==========================
core/excel_export.py 테스트
"""
import pytest
import tempfile
import logging
from pathlib import Path
import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

from core.models import BatchJob, BatchStatus, BatchSummary
from core.excel_export import export_batch_to_excel


log = logging.getLogger(__name__)


def _make_jobs() -> list:
    j1 = BatchJob(1, Path("video1.mp4"))
    j1.status = BatchStatus.DONE

    j2 = BatchJob(2, Path("video2.mp4"))
    j2.status = BatchStatus.FAILED
    j2.error_message = "FFmpeg 오류 발생"

    j3 = BatchJob(3, Path("video3.mp4"))
    j3.status = BatchStatus.SKIPPED

    j4 = BatchJob(4, Path("video4.mp4"))
    j4.status = BatchStatus.CANCELLED
    return [j1, j2, j3, j4]


def test_excel_creates_file():
    with tempfile.TemporaryDirectory() as d:
        jobs    = _make_jobs()
        summary = BatchSummary(total=4, done=1, failed=1, skipped=1, cancelled=1)
        path    = export_batch_to_excel(jobs, summary, Path(d), log)
        assert path.exists()
        assert path.suffix == ".xlsx"


def test_excel_filename_has_timestamp():
    with tempfile.TemporaryDirectory() as d:
        path = export_batch_to_excel([], BatchSummary(), Path(d), log)
        assert "batch_result_" in path.name


def test_excel_readable_by_openpyxl():
    """저장된 Excel 파일이 openpyxl로 읽혀야 함."""
    from openpyxl import load_workbook
    with tempfile.TemporaryDirectory() as d:
        jobs    = _make_jobs()
        summary = BatchSummary(total=4, done=1, failed=1, skipped=1, cancelled=1)
        path    = export_batch_to_excel(jobs, summary, Path(d), log)
        wb      = load_workbook(str(path))
        assert "배치 결과" in wb.sheetnames
        assert "요약" in wb.sheetnames


def test_excel_row_count_matches_jobs():
    """데이터 행 수 = 영상 수."""
    from openpyxl import load_workbook
    with tempfile.TemporaryDirectory() as d:
        jobs    = _make_jobs()
        summary = BatchSummary(total=4)
        path    = export_batch_to_excel(jobs, summary, Path(d), log)
        wb      = load_workbook(str(path))
        ws      = wb["배치 결과"]
        # 행1=타이틀, 행2=헤더, 행3~=데이터
        data_rows = ws.max_row - 2
        assert data_rows == len(jobs)


def test_excel_summary_sheet_success_rate():
    """요약 시트에 성공률 값이 있어야 함."""
    from openpyxl import load_workbook
    with tempfile.TemporaryDirectory() as d:
        summary = BatchSummary(total=10, done=7, failed=3)
        path    = export_batch_to_excel([], summary, Path(d), log)
        wb      = load_workbook(str(path))
        ws      = wb["요약"]
        values  = [ws.cell(row=r, column=2).value for r in range(1, 9)]
        assert "70.0%" in values


def test_excel_empty_jobs():
    """빈 jobs도 오류 없이 저장되어야 함."""
    with tempfile.TemporaryDirectory() as d:
        path = export_batch_to_excel([], BatchSummary(), Path(d), log)
        assert path.exists()


def test_excel_error_message_truncated():
    """긴 에러 메시지가 100자로 잘려야 함."""
    from openpyxl import load_workbook
    with tempfile.TemporaryDirectory() as d:
        job = BatchJob(1, Path("v.mp4"))
        job.status = BatchStatus.FAILED
        job.error_message = "A" * 200
        path = export_batch_to_excel([job], BatchSummary(total=1, failed=1), Path(d), log)
        wb   = load_workbook(str(path))
        ws   = wb["배치 결과"]
        err_cell = ws.cell(row=3, column=6).value
        assert len(err_cell) <= 100


def test_excel_creates_output_dir():
    """output_dir가 없어도 자동 생성되어야 함."""
    with tempfile.TemporaryDirectory() as d:
        new_dir = Path(d) / "deep" / "nested" / "reports"
        path    = export_batch_to_excel([], BatchSummary(), new_dir, log)
        assert new_dir.exists()
        assert path.exists()
