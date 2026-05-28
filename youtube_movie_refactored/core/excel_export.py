"""
core/excel_export.py — V2.2
[NEW] 처리 시간(elapsed_str) 열 실제 데이터로 채움
"""
from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import List

from core.models import BatchJob, BatchStatus, BatchSummary, ProcessingError


COL_HEADERS = ["#", "영상 파일명", "상태", "처리 시간", "최종 영상 경로", "오류 메시지"]
COL_WIDTHS  = [5,   40,           12,    12,           60,              50]

STATUS_FILL = {
    BatchStatus.DONE:      "C6EFCE",
    BatchStatus.FAILED:    "FFC7CE",
    BatchStatus.SKIPPED:   "FFEB9C",
    BatchStatus.CANCELLED: "D9D9D9",
    BatchStatus.PENDING:   "FFFFFF",
    BatchStatus.RUNNING:   "BDD7EE",
}
STATUS_KO = {
    BatchStatus.DONE:      "✅ 완료",
    BatchStatus.FAILED:    "❌ 실패",
    BatchStatus.SKIPPED:   "⏭ 재사용",
    BatchStatus.CANCELLED: "⏹ 취소",
    BatchStatus.PENDING:   "⏳ 대기",
    BatchStatus.RUNNING:   "🔄 처리중",
}


def export_batch_to_excel(
    jobs: List[BatchJob],
    summary: BatchSummary,
    output_dir: Path,
    logger: logging.Logger,
) -> Path:
    """배치 결과를 Excel (.xlsx) 로 저장합니다."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ProcessingError(
            "file_write_error",
            "openpyxl이 설치되지 않았습니다.\n실행: pip install openpyxl"
        ) from exc

    wb  = Workbook()
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── 시트1: 상세 결과 ──────────────────────
    ws = wb.active
    ws.title = "배치 결과"

    # 타이틀
    ws.merge_cells("A1:F1")
    tc = ws["A1"]
    tc.value     = f"📦 배치 처리 결과 — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    tc.font      = Font(bold=True, size=14, color="FFFFFF")
    tc.fill      = PatternFill("solid", fgColor="1F4E79")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # 헤더
    hf = PatternFill("solid", fgColor="2E75B6")
    for ci, (hdr, w) in enumerate(zip(COL_HEADERS, COL_WIDTHS), 1):
        c = ws.cell(row=2, column=ci, value=hdr)
        c.font      = Font(bold=True, color="FFFFFF", size=11)
        c.fill      = hf
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 22

    # 데이터
    for ri, job in enumerate(jobs, 3):
        sk  = STATUS_KO.get(job.status, str(job.status))
        rf  = PatternFill("solid", fgColor=STATUS_FILL.get(job.status, "FFFFFF"))
        fp  = str(job.result.final_video_path) if job.result and job.result.final_video_path else ""

        # [NEW] elapsed_str 실제 데이터
        row_vals = [
            job.job_id,
            job.video_path.name,
            sk,
            job.elapsed_str,          # ← 처리 소요 시간
            fp,
            job.error_message[:100] if job.error_message else "",
        ]
        aligns = ["center","left","center","center","left","left"]

        for ci, (val, align) in enumerate(zip(row_vals, aligns), 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill      = rf
            c.alignment = Alignment(horizontal=align, vertical="center",
                                    wrap_text=(ci in (5, 6)))
            c.border    = border
        ws.row_dimensions[ri].height = 18

    # 평균 처리 시간 행
    elapsed_list = [j.elapsed_sec for j in jobs if j.elapsed_sec is not None]
    if elapsed_list:
        avg_sec = sum(elapsed_list) / len(elapsed_list)
        avg_row = len(jobs) + 3
        ws.cell(avg_row, 1, "평균 처리 시간").font = Font(bold=True)
        ws.cell(avg_row, 4,
                f"{avg_sec:.0f}초" if avg_sec < 60 else f"{int(avg_sec//60)}분 {int(avg_sec%60):02d}초"
                ).font = Font(bold=True, color="1F4E79")

    ws.auto_filter.ref = f"A2:F{len(jobs)+2}"
    ws.freeze_panes    = "A3"

    # ── 시트2: 요약 ───────────────────────────
    ws2 = wb.create_sheet("요약")
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 18

    # 총 소요 시간 계산
    total_sec = sum(j.elapsed_sec or 0 for j in jobs)
    total_str = (f"{total_sec:.0f}초" if total_sec < 60
                 else f"{int(total_sec//60)}분 {int(total_sec%60):02d}초")

    rows = [
        ("항목",             "수치",          "1F4E79", "FFFFFF"),
        ("전체 영상 수",      summary.total,   None,     "000000"),
        ("✅ 성공",           summary.done,    "C6EFCE", "276221"),
        ("❌ 실패",           summary.failed,  "FFC7CE", "9C0006"),
        ("⏭ 재사용 (스킵)",  summary.skipped, "FFEB9C", "9C6500"),
        ("⏹ 취소",           summary.cancelled,"D9D9D9","404040"),
        ("성공률",            f"{summary.success_rate}%", "BDD7EE","1F4E79"),
        ("총 처리 시간",      total_str,       None,     "000000"),
        ("생성 일시",         datetime.now().strftime("%Y-%m-%d %H:%M:%S"), None, "000000"),
    ]
    for r, (label, val, bg, fg) in enumerate(rows, 1):
        for ci, v in enumerate([label, val], 1):
            c = ws2.cell(row=r, column=ci, value=v)
            c.font      = Font(bold=(r==1 or ci==1), color=fg, size=11)
            c.alignment = Alignment(horizontal="center" if ci==2 else "left",
                                    vertical="center")
            c.border    = border
            if bg:
                c.fill = PatternFill("solid", fgColor=bg)
        ws2.row_dimensions[r].height = 22

    # ── 저장 ──────────────────────────────────
    output_dir.mkdir(parents=True, exist_ok=True)
    fname     = f"batch_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    xlsx_path = output_dir / fname
    try:
        wb.save(str(xlsx_path))
    except PermissionError as exc:
        raise ProcessingError("file_write_error",
            f"Excel 저장 권한이 없습니다: {xlsx_path}") from exc

    logger.info("Excel 결과 저장: %s", xlsx_path)
    return xlsx_path
